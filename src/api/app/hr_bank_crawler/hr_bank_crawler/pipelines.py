# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

import os
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), '../..'))

import re
from time import time
from datetime import datetime
from openpyxl import Workbook
from scrapy.exceptions import DropItem
from openpyxl.utils.exceptions import IllegalCharacterError

from utils.logger import get_logger 


ACCEPTABLE_FAILURE_RATE = 0.03
TODAY_DATE= str(datetime.today())[:10]

logger = get_logger()


class HrBankCrawlerPipeline(object):
    
    def __init__(self, config):
        self.config = config
        self.save_file_dir = self.config.get('save_file_dir', './')
        if not os.path.exists(self.save_file_dir):
            os.makedirs(self.save_file_dir)

        timestamp = str(time())
        self.filename = f'DEFAULT_RAW_{TODAY_DATE}_{timestamp}.xlsx'
        self.save_file_path = os.path.join(self.save_file_dir, self.filename)
        
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append([
            '職缺編號',
            '更新日期',
            '網址連結',
            '工作名稱',
            '公司名稱',
            '公司網址',
            '上市上櫃',
            '外商公司',
            '員工人數',
            '公司產業',
            '產業別',
            '工作內容',
            '公司福利',
            '職務類別',
            '工作待遇',
            '工作性質',
            '上班地點',
            '工作行政區',
            '管理責任',
            '出差外派',
            '上班時段',
            '休假制度',
            '可上班日',
            '需求人數',
            '應徵人數',
            '接受身份',
            '工作經歷',
            '學歷要求',
            '科系要求',
            '語文條件',
            '擅長工具',
            '工作技能',
            '具備證照',
            '其他條件',
            '雇用類型',
            '代徵企業',
            'create_date',
            'timestamp',
        ])


    @classmethod
    def from_crawler(cls, crawler):
        # Read Config in settings.py
        return cls(
            config = crawler.settings.get('PROJECT_CONFIG'),
        )

    
    def open_spider(self, spider):
        # Set output filename & directory
        self.filename = f'{spider.prefix_filename}_RAW_{TODAY_DATE}.xlsx'
        self.save_file_path = os.path.join(self.save_file_dir, self.filename)
        
        
    def close_spider(self, spider):
        print('=' * 30)
        print('>>> close spider!!!')
        print('=' * 30)
        crawled_job_num = self.ws.max_row
        expected_job_num = spider.accumulated_job_num
        lower_bound_job_num = int((1 - ACCEPTABLE_FAILURE_RATE) * expected_job_num)
        
        if crawled_job_num >= lower_bound_job_num:
            logger.info(f'[CrawledNum] expect:{expected_job_num}, crawled:{crawled_job_num}, filename:{self.filename}')
            print('>>> saving info...')
            self.wb.save(self.save_file_path)
            
        elif crawled_job_num < expected_job_num:
            logger.error(f'[CrawledNum] expect:{expected_job_num}, crawled:{crawled_job_num}, filename:{self.filename}')
            self.wb.save(self.save_file_path)

    
    def process_item(self, item, spider):
        
        if item['job_id']:
            line = [
                item["job_id"],
                item["update_date"],
                item["job_detail_url"],
                item["job_name"],
                item["company_name"],
                item["company_url"],
                item["listed_company"],
                item["foreign_company"],
                item["staff_num"],
                item["company_industry"],
                item["industry"],
                item["job_description"],
                item["welfare"],
                item["job_category"],
                item["salary"],
                item["job_type"],
                item["location"],
                item["district"],
                item["management"],
                item["dispatch"],
                item["working_hours"],
                item["vacation"],
                item["avalibale_work_date"],
                item["required_num"],
                item["applied_num"],
                item["identity"],
                item["experience"],
                item["education"],
                item["department"],
                item["language"],
                item["tool"],
                item["skill"],
                item["certification"],
                item["other_requirement"],
                item["employment_type"],
                item["agent_company"],
                TODAY_DATE,  # create_date
                str(time()), # timestamp
            ]
            
            
            is_exist = False
            for rownum in range(1, self.ws.max_row+1):
                if item["job_id"] == self.ws.cell(rownum, 1).value:
                    is_exist = True
                    return

            if not is_exist:
                try:
                    self.ws.append(line)
                except IllegalCharacterError:
                    line = self._remove_illegal_character(line)
                    self.ws.append(line)
                    
                if self.ws.max_row % 1000 == 0:
                    self.wb.save(self.save_file_path)
                return item
            
        else:
            return
    
    
    def _remove_illegal_character(self, list_):
        for i in range(len(list_)):
            list_[i] = re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', str(list_[i]))
        return list_


